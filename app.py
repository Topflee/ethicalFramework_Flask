from flask import Flask, request, make_response, jsonify, session
from openpyxl import Workbook, load_workbook
import json
from flask_cors import CORS
app = Flask(__name__)
CORS(app)

@app.route('/get_data', methods=['GET'])
def getData():
    if request.method == 'GET':
        with open('data/Dilemna.json') as json_file:
            data = json.load(json_file)
    return make_response(jsonify(data)), 200

@app.route('/post_response', methods=['POST'])
def postResponse():
    if request.method == 'POST':
        participantId    = request.args.get('participantId')
        questionId       = request.args.get('questionId')
        sliderPos        = request.args.get('sliderPos')
        responseNum      = request.args.get('responseNum')
        wb = load_workbook('responses.xlsx')
        ws = wb.active
        numRows = int(ws['B1'].value)
        for row in ws.iter_rows(min_row=numRows, max_col=4, max_row=numRows):
            count = 1
            for cell in row:
                if count == 1:
                    cell.value = participantId
                elif count == 2:
                    cell.value = questionId
                elif count == 3:
                    cell.value = sliderPos
                elif count == 4:
                    cell.value = responseNum
                else:
                    cell.value = "Reduce max col"
                count += 1
        numRows += 1
        ws['B1'].value = numRows
        wb.save('responses.xlsx')
    responseObject = {
                        'status': 'success',
                        'data': "Data posted"
                    }
    return make_response(jsonify(responseObject)), 200

@app.route('/')
def hello_world():
	return 'Hello, World!'

if __name__ == "__main__":
    app.run()